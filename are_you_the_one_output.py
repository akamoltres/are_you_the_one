
from copy import deepcopy
import openpyxl
import pathlib
import re

def generate_probability_table(total_valid, constraints, pair_count):
    probability_table = []
    probability_table.append([i for i in constraints["men"]])
    probability_table[0].insert(0, "")
    for idx, i in enumerate(pair_count):
        probability_table.append([constraints["women"][idx]])
        for j in i:
            if j == 0:
                probability_table[idx + 1].append("X")
            elif j == total_valid:
                probability_table[idx + 1].append("MATCH")
            else:
                probability = float(j) / float(total_valid) * 100.0
                probability_table[idx + 1].append(f"{probability:.2f}%")
    return probability_table

def get_blank_matchups(probability_table):
    matchups = deepcopy(probability_table)
    for i in range(len(matchups)):
        for j in range(len(matchups[i])):
            if i !=0 and j != 0:
                matchups[i][j] = ""
    return matchups

def write_matchup_table(ws, start_cell, week, phase, table, correct = None):
    for idx_r, i in enumerate(table):
        for idx_c, j, in enumerate(i):
            cell = ws.cell(row = idx_r + start_cell[0], column = idx_c + start_cell[1])
            cell.value = j
            if j == "X":
                cell.font = openpyxl.styles.Font(color = "970006")
                cell.fill = openpyxl.styles.PatternFill(fill_type = "solid", fgColor = "FFC7CE")
            elif j == "?" or re.match(r"^[0-9]+(,[0-9]+)*$", j):
                cell.font = openpyxl.styles.Font(color = "9C6500")
                cell.fill = openpyxl.styles.PatternFill(fill_type = "solid", fgColor = "FFEB9C")
            elif j == "MATCH":
                cell.font = openpyxl.styles.Font(color = "006100")
                cell.fill = openpyxl.styles.PatternFill(fill_type = "solid", fgColor = "C6EFCE")
                
    cell = ws.cell(row = start_cell[0], column = start_cell[1])
    cell.value = f"{week}/{phase}"
    if correct is not None:
        cell.value += f" - {correct}"

def write_week_to_xlsx(ws, start_cell, constraints, week, phase, probability_table, matchup_rollup):
    matchup_table = get_blank_matchups(probability_table)

    # add the truth booth to the week's matchup table
    pairs = constraints[f"week{week}"]["booth"]
    if not isinstance(pairs[0], list):
        pairs = [pairs]
    for p in pairs:
        done = False
        for idx_r, i in enumerate(probability_table):
            for idx_c, j in enumerate(i):
                if [probability_table[idx_r][0], probability_table[0][idx_c]] == p or \
                   [probability_table[0][idx_c], probability_table[idx_r][0]] == p:
                    done = True
                    if j == "X":
                        matchup_table[idx_r][idx_c] = "X"
                    elif j == "MATCH":
                        matchup_table[idx_r][idx_c] = "MATCH"
                    else:
                        assert False
        assert done

    # add the matchup ceremony to the week's matchup table
    if phase == "b":
        for pair in constraints[f"week{week}"]["pairs"]:
            done = False
            for idx_r, i in enumerate(probability_table):
                for idx_c, j in enumerate(i):
                    if [probability_table[idx_r][0], probability_table[0][idx_c]] == pair or \
                       [probability_table[0][idx_c], probability_table[idx_r][0]] == pair:
                        done = True
                        if j == "X":
                            matchup_table[idx_r][idx_c] = "X"
                        elif j == "MATCH":
                            matchup_table[idx_r][idx_c] = "MATCH"
                        else:
                            matchup_table[idx_r][idx_c] = "?"
            assert done

    # write this week's matchup to the excel sheet
    write_matchup_table(ws, start_cell, week, phase, matchup_table, None if phase == "a" else str(constraints[f"week{week}"]["correct"]) + " correct")

    # update matchup rollup
    for idx_r, i in enumerate(matchup_table):
        for idx_c, j in enumerate(i):
            if idx_r != 0 and idx_c != 0 and matchup_table[idx_r][idx_c] != "":
                matchup_rollup[idx_r][idx_c] += (str(week) if matchup_rollup[idx_r][idx_c] == "" else f",{week}")

def write_to_xlsx(filename, total_valid, probability_table, constraints, week, phase):
    # check if the file exists, if it does, ask if it should be overwritten, and if yes, wipe it before continuing
    if pathlib.Path(filename).exists():
        while True:
            print(f"{filename} already exists! Overwrite? [y/N] ", end = "")
            user_input = input()
            if user_input == "" or user_input.lower() == "n" or user_input.lower() == "no":
                return
            elif user_input.lower() == "y" or user_input.lower() == "yes":
                break

    print(f"Writing to {filename}...", end = "")

    # setup
    wb = openpyxl.Workbook()
    ws = wb.active

    # write case count and probability tables
    write_matchup_table(ws, (1, 1), week, phase, probability_table, f"{total_valid} remaining")

    # initialize matchup rollup
    matchup_rollup = get_blank_matchups(probability_table)

    # write the week-by-week matchups
    row_length = 3
    for w in range(1, week + 1):
        write_week_to_xlsx(ws,
                           ((len(probability_table) + 1) * ((w - 1) // row_length + 1) + 1, (len(probability_table[0]) + 1) * ((w - 1) % row_length) + 1),
                           constraints,
                           w,
                           "b" if w != week else phase,
                           probability_table,
                           matchup_rollup)

    # write Xs and MATCHes to the matchup rollup
    for idx_r, i in enumerate(matchup_rollup):
        for idx_c, j in enumerate(i):
            if idx_r != 0 and idx_c != 0 and (probability_table[idx_r][idx_c] == "X" or probability_table[idx_r][idx_c] == "MATCH"):
                matchup_rollup[idx_r][idx_c] = probability_table[idx_r][idx_c]

    # write the matchup rollup
    write_matchup_table(ws, (1, len(probability_table) + 2), week, phase, matchup_rollup)

    wb.save(filename = filename)

    print("complete!")
